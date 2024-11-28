import streamlit as st
import pandas as pd
import openpyxl
import io
import re

class Div7AChecker:
    def __init__(self):
        """
        Initialize the Division 7A Compliance Checker
        """
        self.issues = []
        self.warnings = []
        self.recommendations = []

    def load_workbook(self, uploaded_file):
        """
        Load Excel workbook safely
        """
        try:
            return openpyxl.load_workbook(uploaded_file, data_only=True)
        except Exception as e:
            st.error(f"Error loading workbook: {e}")
            return None

    def check_binder_setup(self, workbook):
        """
        Verify basic binder setup requirements
        """
        # Check for required worksheets
        required_sheets = [
            'D03 Company Debit Loan Summary', 
            'G31 Div 7A Loan Calculator', 
            'D06 Distributable Surplus'
        ]

        missing_sheets = [
            sheet for sheet in required_sheets 
            if sheet not in workbook.sheetnames
        ]

        if missing_sheets:
            self.issues.append(f"Missing critical worksheets: {', '.join(missing_sheets)}")
            self.recommendations.append("Add the missing worksheets to complete Division 7A compliance documentation.")

    def check_loan_details(self, workbook):
        """
        Perform detailed checks on loan details
        """
        # Look for Div 7A Calculator worksheet
        try:
            div7a_sheet = workbook['G31 Div 7A Loan Calculator']
            
            # Check loan characteristics
            loan_type = div7a_sheet['J36'].value  # Assuming location of loan type
            if loan_type not in ['Secured', 'Unsecured']:
                self.warnings.append("Loan type not clearly specified. Recommend clarifying secured/unsecured status.")

            # Verify loan term
            loan_term = div7a_sheet['J39'].value  # Assuming location of loan term
            if loan_term and loan_type == "Secured" and loan_term > 25:
                self.issues.append(f"Loan term for {loan_type} Loan exceeds maximum allowed (25 years). Current term: {loan_term}")
                self.recommendations.append("Adjust loan term or loan type to comply with Division 7A regulations.")
            
            elif loan_term and loan_type == "Unsecured" and loan_term > 7:
                self.issues.append(f"Loan term for {loan_type} Loan exceeds maximum allowed (7 years). Current term: {loan_term}")
                self.recommendations.append("Adjust loan term or loan type to comply with Division 7A regulations.")

        except Exception as e:
            self.issues.append(f"Unable to fully check loan details: {e}")

    def check_dividend_logic(self, workbook):
        """
        Apply simplified logic checks for potential Division 7A dividend
        """
        # This would be a complex check requiring detailed parsing of financial data
        # For now, we'll do basic sanity checks
        try:
            summary_sheet = workbook['D03 Company Debit Loan Summary']
            
            # Look for any indicators of potential dividend scenarios
            loan_balance = summary_sheet['G87'].value if summary_sheet['G87'].value else 0
            
            if loan_balance > 0:
                self.warnings.append(f"Existing loan balance detected: ${loan_balance}")
                self.recommendations.append("Verify if loan meets Division 7A exemption criteria.")

        except Exception as e:
            self.issues.append(f"Unable to perform dividend logic check: {e}")

    def check_minimum_repayments(self, workbook):
        """
        Verify minimum repayment calculations
        """
        try:
            div7a_sheet = workbook['G31 Div 7A Loan Calculator']
            
            # Look for minimum repayment section
            min_repayment = div7a_sheet['J44'].value  # Assuming location of minimum repayment
            
            if min_repayment is None or min_repayment == 0:
                self.warnings.append("No minimum repayment calculated. Verify compliance.")
                self.recommendations.append("Ensure minimum repayment is calculated for each loan year.")

        except Exception as e:
            self.issues.append(f"Unable to verify minimum repayments: {e}")

    def comprehensive_check(self, uploaded_file):
        """
        Perform comprehensive check on uploaded Excel file
        """
        # Reset findings
        self.issues = []
        self.warnings = []
        self.recommendations = []

        # Load workbook
        workbook = self.load_workbook(uploaded_file)
        
        if not workbook:
            return

        # Run all verification checks
        self.check_binder_setup(workbook)
        self.check_loan_details(workbook)
        self.check_dividend_logic(workbook)
        self.check_minimum_repayments(workbook)

        return {
            'issues': self.issues,
            'warnings': self.warnings,
            'recommendations': self.recommendations
        }

def main():
    st.title("Division 7A Compliance Checker")
    
    # Add custom CSS for better styling
    st.markdown("""
    <style>
    .stAlert {
        margin-bottom: 10px;
    }
    .reportview-container {
        background: #f0f2f6;
    }
    </style>
    """, unsafe_allow_html=True)

    # File uploader
    uploaded_file = st.file_uploader(
        "Upload Division 7A Excel File", 
        type=['xlsx', 'xls'],
        help="Upload your Division 7A calculator Excel file for compliance checking"
    )

    # Sidebar for additional information
    st.sidebar.header("Division 7A Compliance Guide")
    st.sidebar.info("""
    ### Key Verification Points
    - Binder setup accuracy
    - Loan term compliance
    - Dividend determination
    - Minimum repayment calculations
    """)

    # Compliance checking
    if uploaded_file is not None:
        # Create checker instance
        checker = Div7AChecker()

        # Perform comprehensive check
        results = checker.comprehensive_check(uploaded_file)

        # Display results
        st.header("Compliance Check Results")

        # Issues (Red - Critical)
        if results['issues']:
            st.error("🚨 Critical Issues Detected")
            for issue in results['issues']:
                st.error(issue)

        # Warnings (Yellow - Needs Attention)
        if results['warnings']:
            st.warning("⚠️ Potential Compliance Warnings")
            for warning in results['warnings']:
                st.warning(warning)

        # Recommendations (Blue - Suggested Actions)
        if results['recommendations']:
            st.info("💡 Recommended Actions")
            for recommendation in results['recommendations']:
                st.info(recommendation)

        # All good message
        if not (results['issues'] or results['warnings'] or results['recommendations']):
            st.success("✅ No compliance issues detected. Document appears to be in order.")

if __name__ == "__main__":
    main()