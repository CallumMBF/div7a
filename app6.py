import streamlit as st
import pandas as pd
import openai
from openai import OpenAI
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook

class LoanAssessmentApp:
    def __init__(self, openai_api_key=None):
        """
        Initialize the Loan Assessment Application
        
        :param openai_api_key: OpenAI API key for LLM analysis
        """
        self.client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        
        # Comprehensive prompt for loan detection
        self.loan_detection_prompt = """
        Analyze the account name: '{account_name}'
        
        Loan Detection Criteria:
        1. Is this account likely to represent a loan?
        2. Provide a clear classification: 'Yes' (definite loan), 'Unsure' (potential loan), or 'No' (not a loan)
        3. If 'Yes' or 'Unsure', explain your reasoning in a brief one sentence answer.
        
        Consider:
        - Explicit loan terms (loan, lending, credit)
        - Financial instrument references
        - Contextual indicators of borrowing
        
        Response Format:
        Classification: [Yes/Unsure/No]
        Reasoning: [Explanation of classification]
        """
    
    def detect_loan(self, account_name):
        """
        Use LLM to detect if an account is a loan
        
        :param account_name: Name of the account to analyze
        :return: Dictionary with loan detection details
        """
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a financial classification expert."},
                    {"role": "user", "content": self.loan_detection_prompt.format(account_name=account_name)}
                ],
                max_tokens=500,
                temperature=0.2
            )
            
            # Parse the response
            result_text = response.choices[0].message.content.strip()
            
            # Extract classification and reasoning
            lines = result_text.split('\n')
            classification = lines[0].split(':')[1].strip() if ':' in lines[0] else 'No'
            reasoning = lines[1].split(':')[1].strip() if len(lines) > 1 and ':' in lines[1] else 'No specific reasoning provided'
            
            return {
                'account_name': account_name,
                'classification': classification,
                'reasoning': reasoning
            }
        
        except Exception as e:
            return {
                'account_name': account_name,
                'classification': 'Error',
                'reasoning': f'Loan Detection Error: {str(e)}'
            }
    
    def determine_loan_date(self, worksheet_path, template_type="unknown"):
        """
        Determine the loan date from the worksheet
        
        :param worksheet_path: Path to the Excel worksheet
        :param template_type: Type of template (Div7A or unknown)
        :return: Dictionary with loan date details
        """
        try:
            # For Div7A template, check cell J37
            if template_type == "Div7A":
                wb = load_workbook(worksheet_path, data_only=True)
                ws = wb.active
                loan_date = ws['J37'].value
                
                return {
                    'date': loan_date,
                    'method': 'Div7A Template',
                    'reasoning': 'Date extracted from standard Div7A template cell J37'
                }
            
            # For other templates, use LLM to analyze
            wb = load_workbook(worksheet_path, data_only=True)
            ws = wb.active
            
            # Convert worksheet data to a format suitable for LLM
            worksheet_data = []
            for row in ws.iter_rows(values_only=True):
                # Filter out None or empty rows
                filtered_row = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                if filtered_row:
                    worksheet_data.append(filtered_row)
            
            # Prepare LLM prompt for date detection
            date_detection_prompt = f"""
            Analyze the following worksheet data to determine the loan origination date:
            
            {worksheet_data[:10]}  # Limit to first 10 rows to avoid token overflow
            
            Guidelines:
            1. Look for dates related to loan origination
            2. Check column headers, first few rows
            3. If nothing identified in column headers or first few rows then check for terms like:
            Income Year of Loan
            Date of Loan
            Year Loan Initiated
            Loan Origination Date
            Loan Disbursement Date
            Loan Start Date
            Loan Agreement Date
            Date Funds Disbursed
            Loan Execution Date
            Date of Loan Issuance
            Date Loan Approved
            Contract Date
            Funding Date
            Date of Advance
            Date Credit Was Extended
            Effective Date of Loan
            Loan Commitment Date
            Principal Issuance Date
            Date Loan Initiated
            Origination Date
            Agreement Signed Date
            Date Note Was Executed
            Advance Date
            Loan Drawdown Date
            Agreement Execution Date
            Lending Date
            Date Collateral Secured
            Financial Closing Date
            Initial Loan Payment Date
            Loan Opening Date
            Contract Initiation Date
            Promissory Note Date
            Approval to Fund Date
            Loan Document Signing Date
            First Funding Date
            Date Financial Terms Were Agreed
            Underwriting Completion Date
            Offer Acceptance Date (Loan)
            Bank Disbursement Date
            Credit Facility Initiation Date
            Closing Date of Loan Agreement
            Date Principal Advanced
            Payment Schedule Start Date
            Borrowing Date
            Origination Agreement Date
            Commitment Effective Date
            Loan Offer Date
            Transaction Date
            Documented Date of Loan
            Approval Date of Lending Terms
            4. The date may be in one of the cells adjacent to the cell with the date name from step 3
            5. Identify the most likely loan start date
        6. Provide your reasoning with a short one sentence answer
            
            Response Format:
            Date: [YYYY-MM-DD or YYYY]
            Reasoning: [Detailed explanation of how you determined the date]
            """
            
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert in financial document analysis."},
                    {"role": "user", "content": date_detection_prompt}
                ],
                max_tokens=500,
                temperature=0.3
            )
            
            # Parse LLM response
            result_text = response.choices[0].message.content.strip()
            lines = result_text.split('\n')
            
            # Extract date and reasoning
            date_line = [line for line in lines if line.startswith('Date:')][0] if lines else None
            reasoning_line = [line for line in lines if line.startswith('Reasoning:')][0] if lines else None
            
            loan_date = date_line.split(':')[1].strip() if date_line else None
            reasoning = reasoning_line.split(':')[1].strip() if reasoning_line else 'No specific reasoning provided'
            
            return {
                'date': loan_date,
                'method': 'LLM Analysis',
                'reasoning': reasoning
            }
        
        except Exception as e:
            return {
                'date': None,
                'method': 'Error',
                'reasoning': f'Loan Date Detection Error: {str(e)}'
            }
    
    def is_loan_older_than_12_months(self, loan_date, financial_year_end):
        """
        Check if loan is older than 12 months
        
        :param loan_date: Date of loan origination
        :param financial_year_end: End of financial year
        :return: Boolean indicating if loan is older than 12 months
        """
        try:
            # Handle different date formats
            if isinstance(loan_date, str):
                # Try parsing different date formats
                date_formats = ['%Y-%m-%d', '%Y-%m', '%Y']
                parsed_date = None
                
                for date_format in date_formats:
                    try:
                        parsed_date = datetime.strptime(loan_date, date_format)
                        
                        # If only year is provided, set to beginning of that year
                        if date_format == '%Y':
                            parsed_date = datetime(parsed_date.year, 1, 1)
                        
                        break
                    except ValueError:
                        continue
                
                if not parsed_date:
                    return True  # Default to True if date can't be parsed (conservative approach)
            elif isinstance(loan_date, datetime):
                parsed_date = loan_date
            else:
                return True  # Default to True if date is not recognizable
            
            # If financial_year_end is a date object, use it directly
            if isinstance(financial_year_end, datetime):
                year_end = financial_year_end
            else:
                # Convert to datetime if it's not already
                year_end = datetime.combine(financial_year_end, datetime.min.time())
            
            # Calculate if loan is older than 12 months
            return (year_end - parsed_date).days > 365
        
        except Exception:
            return True  # Conservative default if any error occurs

def analyze_account(loan_app, account_name, worksheets_directory, financial_year_end):
    """
    Comprehensive analysis for a single account with refined loan detection workflow
    """
    # Initialize result dictionary
    result = {
        'Account': account_name,
        'Loan Detection': {
            'Classification': None,
            'Probability': None,
            'Reasoning': None
        },
        'Loan Date': None,
        'Loan Age': None,
        'Summary': None,
        'Div7A Status': 'Requires further investigation'
    }

    # Detect if account is a loan
    loan_detection = loan_app.detect_loan(account_name)
    
    # Store loan detection details
    result['Loan Detection']['Classification'] = loan_detection['classification']
    result['Loan Detection']['Reasoning'] = loan_detection['reasoning']
    
    # If not a loan
    if loan_detection['classification'] == 'No':
        result['Summary'] = 'Does not appear to be a loan'
        result['Loan Detection']['Probability'] = 'Low'
        result['Div7A Status'] = 'Not applicable'
        return result
    
    # If potentially a loan
    if loan_detection['classification'] in ['Yes', 'Unsure']:
        # Find associated worksheet
        matching_worksheets = [
            f for f in os.listdir(worksheets_directory) 
            if account_name.replace(' ', '_') in f
        ]
        
        # Set probability based on classification
        result['Loan Detection']['Probability'] = 'High' if loan_detection['classification'] == 'Yes' else 'Medium'
        
        if matching_worksheets:
            worksheet_path = os.path.join(worksheets_directory, matching_worksheets[0])
            
            # Determine template type (Div7A or not)
            template_type = "Div7A" if "Div7A" in matching_worksheets[0] else "unknown"
            
            # Get loan date
            loan_date_info = loan_app.determine_loan_date(worksheet_path, template_type)
            
            # Store loan date
            result['Loan Date'] = loan_date_info['date']
            
            # Check loan age
            is_old_loan = loan_app.is_loan_older_than_12_months(
                loan_date_info['date'], 
                financial_year_end
            )
            
            # Determine summary and Div7A requirement
            if is_old_loan is True:
                result['Loan Age'] = 'Older than 12 months'
                result['Summary'] = 'Loan is older than 12 months from the end of financial year'
                result['Div7A Status'] = 'A Div7A may be required if loan is outstanding'
            elif is_old_loan is False:
                result['Loan Age'] = 'Less than 12 months'
                result['Summary'] = 'Loan is not older than 12 months from the end of financial year'
                result['Div7A Status'] = 'Not required in current financial year'
            else:
                result['Loan Age'] = 'Unable to determine'
                result['Summary'] = 'Unable to definitively determine loan age'
                result['Div7A Status'] = 'A Div7A may be required if loan is outstanding and over 12 months'
        else:
            result['Summary'] = 'No matching worksheet found'
            result['Div7A Status'] = 'Requires further investigation'
    
    return result

def main():
    st.title("Div 7a Rec Assessment Tool")
    
    # Sidebar for configuration
    st.sidebar.header("Configuration")
    openai_api_key = os.getenv('OPENAI_API_KEY')

    # File uploaders
    trial_balance_file = st.sidebar.file_uploader("Upload Trial Balance", type=['xlsx'])
    worksheets_directory = st.sidebar.text_input("Worksheets Directory Path")
    financial_year_end = st.sidebar.date_input(
        "Financial Year End", 
        value=datetime(datetime.now().year, 6, 30)
    )
    
    # Initialize results to an empty list
    results = []
    
    # Main analysis
    if st.sidebar.button("Analyze Loans") and trial_balance_file and worksheets_directory and openai_api_key:
        # Initialize Loan Assessment App
        loan_app = LoanAssessmentApp(openai_api_key)
        
        # Read Trial Balance
        trial_balance_df = pd.read_excel(trial_balance_file)
        
        # Progress bar
        progress_bar = st.progress(0)
        
        # Analyze each account
        for index, row in trial_balance_df.iterrows():
            # Update progress
            progress_bar.progress(int((index + 1) / len(trial_balance_df) * 100))
            
            account_name = row['Account Name']
            
            # Perform comprehensive account analysis
            current_dir = os.getcwd()
            worksheets_directory = os.path.join("C:/Users/CallumMatchett/python/streamlit/div7a", worksheets_directory)
            account_result = analyze_account(
                loan_app, 
                account_name, 
                worksheets_directory, 
                financial_year_end
            )
            
            results.append(account_result)
    
    # Display results section
    if results:
        # Create DataFrame for easier display
        results_df = pd.DataFrame(results)
        
        st.header("Loan Assessment Results")
        
        for _, result in results_df.iterrows():
            # Determine color based on summary
            if result['Summary'] == 'Does not appear to be a loan':
                status_color = 'green'
            elif result['Div7A Status'] == 'A Div7A may be required if loan is outstanding':
                status_color = 'red'
            else:
                status_color = 'orange'
            
            # Create expandable section for each account
            with st.expander(f"{result['Account']} - {result['Summary']}", expanded=False):
                # Loan Detection Details
                st.markdown(f"**Summary:** <font color='{status_color}'>{result['Summary']}</font>", unsafe_allow_html=True)
                
                st.subheader("Loan Detection")
                st.write(f"**Classification:** {result['Loan Detection']['Classification']}")
                st.write(f"**Probability:** {result['Loan Detection']['Probability']}")
                st.write(f"**Reasoning:** {result['Loan Detection']['Reasoning']}")
                
                # Additional Loan Details (if applicable)
                if result['Loan Date']:
                    st.subheader("Loan Details")
                    st.write(f"**Loan Date:** {result['Loan Date']}")
                    st.write(f"**Loan Age:** {result['Loan Age']}")
                    st.write(f"**Div7A Status:** {result['Div7A Status']}")
        
        # Download option for full results
        st.download_button(
            label="Download Full Results",
            data=results_df.to_csv(index=False),
            file_name="loan_assessment_results.csv",
            mime="text/csv"
        )
            
    else:
        st.warning("No accounts analyzed")

if __name__ == "__main__":
    main()