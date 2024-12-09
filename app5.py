import streamlit as st
import pandas as pd
import openai
from openai import OpenAI
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook

class LoanAssessmentApp:
    def __init__(self, openai_api_key=None):
        """
        Initialize the Loan Assessment Application
        
        :param openai_api_key: OpenAI API key for LLM analysis
        """
        openai_api_key = os.getenv('OPENAI_API_KEY')
        self.client = OpenAI(api_key=openai_api_key or os.getenv('OPENAI_API_KEY'))
        client = OpenAI(
            api_key=os.getenv("OPENAI_API_KEY"),  # This is the default and can be omitted
        )

        # Comprehensive prompt for loan detection
        self.loan_detection_prompt = """
        Analyze the account name: '{account_name}'
        
        Loan Detection Criteria:
        1. Is this account likely to represent a loan?
        2. Provide a clear classification: 'Yes' (definite loan), 'Unsure' (potential loan), or 'No' (not a loan)
        3. If 'Yes' or 'Unsure', explain your reasoning
        
        Consider:
        - Explicit loan terms (loan, lending, credit)
        - Financial instrument references
        - Contextual indicators of borrowing
        
        Response Format:
        Classification: [Yes/Unsure/No]
        Reasoning: [Explanation of classification]
        """
    
    def detect_loan(self, account_name):
        """
        Use LLM to detect if an account is a loan
        
        :param account_name: Name of the account to analyze
        :return: Dictionary with loan detection details
        """
        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a financial classification expert."},
                    {"role": "user", "content": self.loan_detection_prompt.format(account_name=account_name)}
                ],
                max_tokens=150,
                temperature=0.2
            )
            
            # Parse the response
            result_text = response.choices[0].message.content.strip()
            
            # Extract classification and reasoning
            lines = result_text.split('\n')
            classification = lines[0].split(':')[1].strip() if ':' in lines[0] else 'No'
            reasoning = lines[1].split(':')[1].strip() if len(lines) > 1 and ':' in lines[1] else 'No specific reasoning provided'
            
            return {
                'account_name': account_name,
                'classification': classification,
                'reasoning': reasoning
            }
        
        except Exception as e:
            st.error(f"Loan Detection Error for {account_name}: {e}")
            return {
                'account_name': account_name,
                'classification': 'Unsure',
                'reasoning': f'Error in analysis: {str(e)}'
            }
    
    def determine_loan_date(self, worksheet_path, template_type="unknown"):
        """
        Determine the loan date from the worksheet
        
        :param worksheet_path: Path to the Excel worksheet
        :param template_type: Type of template (Div7A or unknown)
        :return: Dictionary with loan date details
        """
        try:
            # For Div7A template, check cell J37
            if template_type == "Div7A":
                wb = load_workbook(worksheet_path, data_only=True)
                ws = wb.active
                loan_date = ws['J37'].value
                
                return {
                    'date': loan_date,
                    'method': 'Div7A Template',
                    'reasoning': 'Date extracted from standard Div7A template cell J37'
                }
            
            # For other templates, use LLM to analyze
            wb = load_workbook(worksheet_path, data_only=True)
            ws = wb.active
            
            # Convert worksheet data to a format suitable for LLM
            worksheet_data = []
            for row in ws.iter_rows(values_only=True):
                # Filter out None or empty rows
                filtered_row = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                if filtered_row:
                    worksheet_data.append(filtered_row)
            
            # Prepare LLM prompt for date detection
            date_detection_prompt = f"""
            Analyze the following worksheet data to determine the loan origination date:
            
            {worksheet_data[:10]}  # Limit to first 10 rows to avoid token overflow
            
            Guidelines:
            1. Look for dates related to loan origination
            2. Check column headers, first few rows
            3. Identify the most likely loan start date
            4. Provide your reasoning
            
            Response Format:
            Date: [YYYY-MM-DD or YYYY]
            Reasoning: [Detailed explanation of how you determined the date]
            """
            
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert in financial document analysis."},
                    {"role": "user", "content": date_detection_prompt}
                ],
                max_tokens=250,
                temperature=0.3
            )
            
            # Parse LLM response
            result_text = response.choices[0].message.content.strip()
            lines = result_text.split('\n')
            
            # Extract date and reasoning
            date_line = [line for line in lines if line.startswith('Date:')][0] if lines else None
            reasoning_line = [line for line in lines if line.startswith('Reasoning:')][0] if lines else None
            
            loan_date = date_line.split(':')[1].strip() if date_line else None
            reasoning = reasoning_line.split(':')[1].strip() if reasoning_line else 'No specific reasoning provided'
            
            return {
                'date': loan_date,
                'method': 'LLM Analysis',
                'reasoning': reasoning
            }
        
        except Exception as e:
            st.error(f"Loan Date Detection Error: {e}")
            return {
                'date': None,
                'method': 'Error',
                'reasoning': f'Unable to determine loan date: {str(e)}'
            }
    
    def is_loan_older_than_12_months(self, loan_date, financial_year_end):
        """
        Check if loan is older than 12 months
        
        :param loan_date: Date of loan origination
        :param financial_year_end: End of financial year
        :return: Boolean indicating if loan is older than 12 months
        """
        try:
            # Handle different date formats
            if isinstance(loan_date, str):
                # Try parsing different date formats
                date_formats = ['%Y-%m-%d', '%Y-%m', '%Y']
                parsed_date = None
                
                for date_format in date_formats:
                    try:
                        parsed_date = datetime.strptime(loan_date, date_format)
                        break
                    except ValueError:
                        continue
                
                if not parsed_date:
                    st.warning(f"Could not parse loan date: {loan_date}")
                    return False
            elif isinstance(loan_date, datetime):
                parsed_date = loan_date
            else:
                st.warning(f"Unexpected loan date format: {loan_date}")
                return False
            
            # Calculate if loan is older than 12 months
            return (financial_year_end - parsed_date).days > 365
        
        except Exception as e:
            st.error(f"Error checking loan age: {e}")
            return False

def main():
    st.title("Advanced Loan Assessment Tool")
    
    # Sidebar for configuration
    st.sidebar.header("Configuration")
    # openai_api_key = st.sidebar.text_input("OpenAI API Key", type="password")
    openai_api_key = os.getenv('OPENAI_API_KEY')

    # File uploaders
    trial_balance_file = st.sidebar.file_uploader("Upload Trial Balance", type=['xlsx'])
    worksheets_directory = st.sidebar.text_input("Worksheets Directory Path")
    financial_year_end = st.sidebar.date_input(
        "Financial Year End", 
        value=datetime(datetime.now().year, 6, 30)
    )
    
    # Main analysis
    if st.sidebar.button("Analyze Loans") and trial_balance_file and worksheets_directory and openai_api_key:
        # Initialize Loan Assessment App
        loan_app = LoanAssessmentApp(openai_api_key)
        
        # Read Trial Balance
        trial_balance_df = pd.read_excel(trial_balance_file)
        
        # Results container
        results = []
        
        # Progress bar
        progress_bar = st.progress(0)
        
        # Analyze each account
        for index, row in trial_balance_df.iterrows():
            # Update progress
            progress_bar.progress(int((index + 1) / len(trial_balance_df) * 100))
            
            account_name = row['Account Name']
            
            # Detect if account is a loan
            loan_detection = loan_app.detect_loan(account_name)
            
            # If loan or potentially a loan
            if loan_detection['classification'] in ['Yes', 'Unsure']:
                # Find associated worksheet
                current_dir = os.getcwd()
                worksheets_directory = os.path.join("C:/Users/CallumMatchett/python/streamlit/div7a", worksheets_directory)
                matching_worksheets = [

                    f for f in os.listdir(worksheets_directory) 
                    if account_name.replace(' ', '_') in f
                ]
                
                if matching_worksheets:
                    worksheet_path = os.path.join(worksheets_directory, matching_worksheets[0])
                    
                    # Determine template type (Div7A or not)
                    template_type = "Div7A" if "Div7A" in matching_worksheets[0] else "unknown"
                    
                    # Get loan date
                    loan_date_info = loan_app.determine_loan_date(worksheet_path, template_type)
                    
                    # Check loan age
                    if loan_date_info['date']:
                        is_old_loan = loan_app.is_loan_older_than_12_months(
                            loan_date_info['date'], 
                            financial_year_end
                        )
                        
                        result = {
                            'Account': account_name,
                            'Loan Detection': loan_detection['classification'],
                            'Detection Reasoning': loan_detection['reasoning'],
                            'Loan Date': loan_date_info['date'],
                            'Date Detection Method': loan_date_info['method'],
                            'Date Detection Reasoning': loan_date_info['reasoning'],
                            'Older Than 12 Months': is_old_loan
                        }
                        
                        results.append(result)
        
        # Display results
        if results:
            results_df = pd.DataFrame(results)
            st.dataframe(results_df)
            
            # Download option
            st.download_button(
                label="Download Results",
                data=results_df.to_csv(index=False),
                file_name="loan_assessment_results.csv",
                mime="text/csv"
            )
        else:
            st.warning("No loans detected")

if __name__ == "__main__":
    main()