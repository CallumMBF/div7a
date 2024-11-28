import streamlit as st
import pandas as pd
import openai
from openai import OpenAI
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook

class LoanAssessmentApp:
    def __init__(self, openai_api_key=None):
        """Initialize the Loan Assessment Application"""
        self.client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        self.loan_detection_prompt = """
        Analyze the account name: '{account_name}'
        
        Loan Detection Criteria:
        1. Is this account likely to represent a loan?
        2. Provide a clear classification: 'Yes' (definite loan), 'Unsure' (potential loan), or 'No' (not a loan)
        3. If 'Yes' or 'Unsure', explain your reasoning in a brief one sentence answer.
        
        Consider:
        - Explicit loan terms (loan, lending, credit)
        - Financial instrument references
        - Contextual indicators of borrowing
        
        Response Format:
        Classification: [Yes/Unsure/No]
        Reasoning: [Explanation of classification]
        """
        
        # Initial seed list of date keywords - will be expanded dynamically
        self.date_keywords = [
            "Year Loan Initiated",
            "Date of Loan",
            "Loan Start Date",
            "Loan Agreement Date",
        ]

    def analyze_cell_content(self, cell_text):
        """
        Use LLM to determine if a cell might contain a loan date reference
        """
        date_analysis_prompt = """
        You are a financial expert specializing in loan documentation. Analyze the following cell text:

        "{cell_text}"

        Determine if this text appears to be referring to a loan initiation date, start date, or similar concept.
        Consider variations and alternative phrasings that mean the same thing.
        
        Examples of what to look for:
        - Direct references to loan dates (e.g., "Loan Date", "Date of Loan")
        - Temporal markers related to loan initiation (e.g., "Started", "Commenced", "Initiated")
        - References to loan documentation timing (e.g., "Signed", "Executed", "Documented")
        - Any phrase that could reasonably indicate when a loan began
        
        Response Format:
        Is Date Reference: [Yes/No]
        Confidence: [High/Medium/Low]
        Reasoning: [Brief explanation]
        """

        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a financial expert focusing on loan documentation analysis."},
                    {"role": "user", "content": date_analysis_prompt.format(cell_text=cell_text)}
                ],
                max_tokens=150,
                temperature=0.2
            )
            
            result_text = response.choices[0].message.content.strip()
            lines = result_text.split('\n')
            
            is_date_ref = lines[0].split(':')[1].strip() if len(lines) > 0 else 'No'
            confidence = lines[1].split(':')[1].strip() if len(lines) > 1 else 'Low'
            
            return is_date_ref == 'Yes' and confidence in ['High', 'Medium']
            
        except Exception:
            return False

    def detect_loan(self, account_name):
        """Use LLM to detect if an account is a loan"""
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a financial classification expert."},
                    {"role": "user", "content": self.loan_detection_prompt.format(account_name=account_name)}
                ],
                max_tokens=500,
                temperature=0.2
            )
            
            result_text = response.choices[0].message.content.strip()
            lines = result_text.split('\n')
            classification = lines[0].split(':')[1].strip() if ':' in lines[0] else 'No'
            reasoning = lines[1].split(':')[1].strip() if len(lines) > 1 and ':' in lines[1] else 'No specific reasoning provided'
            
            return {
                'account_name': account_name,
                'classification': classification,
                'reasoning': reasoning
            }
        except Exception as e:
            return {
                'account_name': account_name,
                'classification': 'Error',
                'reasoning': f'Loan Detection Error: {str(e)}'
            }

    def find_date_in_worksheet(self, ws):
        """
        Search through worksheet to find cells containing date references and their adjacent date values
        """
        date_info = []
        
        # Extended search range
        search_rows = 1000  # Search up to 1000 rows
        search_cols = 50    # Search up to 50 columns
        
        for row in range(1, search_rows + 1):
            for col in range(1, search_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value).strip() if cell.value else ""
                
                if not cell_value:
                    continue
                
                # Check if cell contains a date keyword
                is_date_reference = any(keyword.lower() in cell_value.lower() 
                                    for keyword in self.date_keywords)
                
                if is_date_reference:
                    # Check adjacent cells for dates
                    adjacent_cells = [
                        (row, col + 1, "Right"),
                        (row + 1, col, "Below"),
                        (row - 1, col, "Above"),
                        (row, col - 1, "Left")
                    ]
                    
                    for adj_row, adj_col, direction in adjacent_cells:
                        if 1 <= adj_row <= search_rows and 1 <= adj_col <= search_cols:
                            adj_cell = ws.cell(row=adj_row, column=adj_col)
                            adj_value = adj_cell.value
                            
                            # Check if adjacent cell contains a date
                            if isinstance(adj_value, (datetime, pd.Timestamp)):
                                date_info.append({
                                    'date': adj_value,
                                    'keyword': cell_value,
                                    'location': f'Cell {adj_cell.coordinate}',
                                    'context': f'Found {direction} of "{cell_value}" in cell {cell.coordinate}'
                                })
                            # Check for year as integer or string
                            elif isinstance(adj_value, (int, str)):
                                str_value = str(adj_value).strip()
                                if str_value.isdigit():
                                    year = int(str_value)
                                    if 1900 <= year <= 2100:  # Reasonable year range
                                        date_info.append({
                                            'date': datetime(year, 1, 1),
                                            'keyword': cell_value,
                                            'location': f'Cell {adj_cell.coordinate}',
                                            'context': f'Found year {direction} of "{cell_value}" in cell {cell.coordinate}'
                                        })
        
        return date_info

    def determine_loan_date(self, worksheet_path, template_type="unknown"):
        """Determine the loan date from the worksheet"""
        try:
            # For Div7A template, check cell J37
            if template_type == "Div7A":
                wb = load_workbook(worksheet_path, data_only=True)
                ws = wb.active
                loan_date = ws['J37'].value
                
                if loan_date:
                    return {
                        'date': loan_date,
                        'method': 'Div7A Template',
                        'reasoning': 'Date extracted from standard Div7A template cell J37'
                    }
            
            # For all templates, search for date keywords and analyze cells
            wb = load_workbook(worksheet_path, data_only=True)
            ws = wb.active
            
            date_findings = self.find_date_in_worksheet(ws)
            
            if date_findings:
                # Sort findings by date (earliest first) and confidence
                date_findings.sort(key=lambda x: x['date'])
                best_finding = date_findings[0]
                
                return {
                    'date': best_finding['date'],
                    'method': 'Cell Analysis',
                    'reasoning': best_finding['context']
                }
            
            return {
                'date': None,
                'method': 'Not Found',
                'reasoning': 'No loan date could be determined from the worksheet'
            }
            
        except Exception as e:
            return {
                'date': None,
                'method': 'Error',
                'reasoning': f'Loan Date Detection Error: {str(e)}'
            }

    def is_loan_older_than_12_months(self, loan_date, financial_year_end):
        """Check if loan is older than 12 months"""
        try:
            if isinstance(loan_date, str):
                date_formats = ['%Y-%m-%d', '%Y-%m', '%Y']
                parsed_date = None
                
                for date_format in date_formats:
                    try:
                        parsed_date = datetime.strptime(loan_date, date_format)
                        if date_format == '%Y':
                            parsed_date = datetime(parsed_date.year, 1, 1)
                        break
                    except ValueError:
                        continue
                
                if not parsed_date:
                    return True
            elif isinstance(loan_date, datetime):
                parsed_date = loan_date
            else:
                return True
            
            if isinstance(financial_year_end, datetime):
                year_end = financial_year_end
            else:
                year_end = datetime.combine(financial_year_end, datetime.min.time())
            
            return (year_end - parsed_date).days > 365
        
        except Exception:
            return True

def analyze_account(loan_app, account_name, current_year_total, worksheet_link, worksheets_directory, financial_year_end):
    """Comprehensive analysis for a single account"""
    result = {
        'Account': account_name,
        'Current Year Total': current_year_total,
        'Loan Detection': {
            'Classification': None,
            'Probability': None,
            'Reasoning': None
        },
        'Loan Date': None,
        'Loan Age': None,
        'Summary': None,
        'Div7A Status': 'Requires further investigation'
    }

    loan_detection = loan_app.detect_loan(account_name)
    
    result['Loan Detection']['Classification'] = loan_detection['classification']
    result['Loan Detection']['Reasoning'] = loan_detection['reasoning']
    
    if loan_detection['classification'] == 'No':
        result['Summary'] = 'Does not appear to be a loan'
        result['Loan Detection']['Probability'] = 'Low'
        result['Div7A Status'] = 'Not applicable'
        return result
    
    if loan_detection['classification'] in ['Yes', 'Unsure']:
        result['Loan Detection']['Probability'] = 'High' if loan_detection['classification'] == 'Yes' else 'Medium'
        
        if worksheet_link:
            # Construct the full path to the worksheet
            worksheet_path = os.path.join(worksheets_directory, os.path.basename(worksheet_link))
            
            if os.path.exists(worksheet_path):
                template_type = "Div7A" if "Div7A" in worksheet_path else "unknown"
                
                loan_date_info = loan_app.determine_loan_date(worksheet_path, template_type)
                result['Loan Date'] = loan_date_info['date']
                
                is_old_loan = loan_app.is_loan_older_than_12_months(
                    loan_date_info['date'], 
                    financial_year_end
                )
                
                if is_old_loan is True:
                    result['Loan Age'] = 'Older than 12 months'
                    result['Summary'] = 'Loan is older than 12 months from the end of financial year'
                    result['Div7A Status'] = 'A Div7A may be required if loan is outstanding'
                elif is_old_loan is False:
                    result['Loan Age'] = 'Less than 12 months'
                    result['Summary'] = 'Loan is not older than 12 months from the end of financial year'
                    result['Div7A Status'] = 'Not required in current financial year'
                else:
                    result['Loan Age'] = 'Unable to determine'
                    result['Summary'] = 'Unable to definitively determine loan age'
                    result['Div7A Status'] = 'A Div7A may be required if loan is outstanding and over 12 months'
            else:
                result['Summary'] = f'Worksheet not found at path: {worksheet_path}'
                result['Div7A Status'] = 'Requires further investigation'
        else:
            result['Summary'] = 'No worksheet link provided'
            result['Div7A Status'] = 'Requires further investigation'
    
    return result

def main():
    st.title("Div 7a Rec Assessment Tool")
    
    st.sidebar.header("Configuration")
    openai_api_key = os.getenv('OPENAI_API_KEY')

    trial_balance_file = st.sidebar.file_uploader("Upload Trial Balance", type=['xlsx'])
    worksheets_directory = st.sidebar.text_input("Worksheets Directory Path")
    financial_year_end = st.sidebar.date_input(
        "Financial Year End", 
        value=datetime(2023, 6, 30)
    )
    
    results = []
    
    if st.sidebar.button("Analyze Accounts") and trial_balance_file and worksheets_directory and openai_api_key:
        loan_app = LoanAssessmentApp(openai_api_key)
        trial_balance_df = pd.read_excel(trial_balance_file)
        
        # Verify required columns exist
        required_columns = ['Account Name', 'Current Year Total', 'Worksheet Link']
        missing_columns = [col for col in required_columns if col not in trial_balance_df.columns]
        
        if missing_columns:
            st.error(f"Missing required columns in trial balance: {', '.join(missing_columns)}")
            return
        
        progress_bar = st.progress(0)
        
        for index, row in trial_balance_df.iterrows():
            progress_bar.progress(int((index + 1) / len(trial_balance_df) * 100))
            account_result = analyze_account(
                loan_app, 
                row['Account Name'],
                row['Current Year Total'],
                row['Worksheet Link'],
                worksheets_directory, 
                financial_year_end
            )
            results.append(account_result)
    
    if results:
        results_df = pd.DataFrame(results)
        st.header("Loan Assessment Results")
        
        for _, result in results_df.iterrows():
            if result['Summary'] == 'Does not appear to be a loan':
                status_color = 'green'
            elif result['Div7A Status'] == 'A Div7A may be required if loan is outstanding':
                status_color = 'red'
            else:
                status_color = 'orange'
            
            with st.expander(f"{result['Account']} - {result['Summary']}", expanded=False):
                st.markdown(f"**Summary:** <font color='{status_color}'>{result['Summary']}</font>", unsafe_allow_html=True)
                
                st.subheader("Loan Detection")
                st.write(f"**Classification:** {result['Loan Detection']['Classification']}")
                st.write(f"**Probability:** {result['Loan Detection']['Probability']}")
                st.write(f"**Reasoning:** {result['Loan Detection']['Reasoning']}")
                
                if result['Loan Date']:
                    st.subheader("Loan Details")
                    st.write(f"**Loan Date:** {result['Loan Date']}")
                    st.write(f"**Loan Age:** {result['Loan Age']}")
                    st.write(f"**Div7A Status:** {result['Div7A Status']}")

        # Download option for full results
        st.download_button(
            label="Download Full Results",
            data=results_df.to_csv(index=False),
            file_name="loan_assessment_results.csv",
            mime="text/csv"
        )
            
    else:
        st.warning("No accounts analyzed")

if __name__ == "__main__":
    main()